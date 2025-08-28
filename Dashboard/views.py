from django.shortcuts import render
from django.http import HttpResponse

from rest_framework import viewsets,generics
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from .serializers import *
from Application.models import *


class DashboardTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class DashboardTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class VehicleCategoryViewSetDashboard(viewsets.ModelViewSet):
    queryset = VehicleCategory.objects.all()
    serializer_class = VehicleCategorySerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class VehicleViewSetDashboard(viewsets.ModelViewSet):
    queryset = Vehicle.objects.all()
    serializer_class = VehicleSerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class BookingViewSetDashboard(viewsets.ModelViewSet):
    queryset = Booking.objects.all()
    serializer_class = BookingSerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class EnquiryBookingViewSetDashboard(viewsets.ModelViewSet):
    queryset = EnquiryBooking.objects.all()
    serializer_class = EnquiryBookingSerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class ProjectGalleryViewSetDashboard(viewsets.ModelViewSet):
    queryset = ProjectGallery.objects.all()
    serializer_class = ProjectGallerySerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class ServicesViewSetDashboard(viewsets.ModelViewSet):
    queryset = Services.objects.all()
    serializer_class = ServicesSerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class HomePageSliderImageViewSetDashboard(viewsets.ModelViewSet):
    queryset = HomePageSliderImage.objects.all()
    serializer_class = HomePageSliderImageSerializerDashboard
    permission_classes = [IsSuperuserOrStaff]

class AboutUsImagesViewSetDashboard(viewsets.ModelViewSet):
    queryset = AboutUsImages.objects.all()
    serializer_class = AboutUsImagesSerializerDashboard
    permission_classes = [IsSuperuserOrStaff]
    
    
class ServiceEnquiryViewSetDashboard(viewsets.ModelViewSet):
    queryset = ServiceEnquiry.objects.all()
    serializer_class = ServiceEnquirySerializerDashboard
    permission_classes = [IsSuperuserOrStaff]
    
    
    
class UpdateEnquiryStatusView(generics.UpdateAPIView):
    queryset = EnquiryBooking.objects.all()
    serializer_class = EnquiryBookingSerializerDashboard
    permission_classes = [IsAuthenticated]  
    lookup_field = "pk"

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)
    
    
class UpdateServiceStatusView(generics.UpdateAPIView):
    queryset = ServiceEnquiry.objects.all()
    serializer_class = ServiceEnquirySerializerDashboard
    permission_classes = [IsAuthenticated]  
    lookup_field = "pk"

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)
    
    
class UpdateBookingStatusView(generics.UpdateAPIView):
    queryset = Booking.objects.all()
    serializer_class = BookingSerializerDashboard
    permission_classes = [IsAuthenticated]  
    lookup_field = "pk"

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)


class ExportBookingsExcel(APIView):
    permission_classes = [IsAuthenticated]  

    def get(self, request, *args, **kwargs):
        filter_date = request.query_params.get("filter")

        bookings = Booking.objects.all().order_by("-date")

        from datetime import date
        today = date.today()

        if filter_date == "today":
            bookings = bookings.filter(date=today)
        elif filter_date == "month":
            bookings = bookings.filter(date__month=today.month, date__year=today.year)
        elif filter_date == "year":
            bookings = bookings.filter(date__year=today.year)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"

        headers = ["Name", "Email", "Title", "Duration", "Date", "Time", "Status"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for b in bookings:
            status = "completed" if b.date < today else "upcoming"
            ws.append([
                b.name,
                b.email,
                b.title,
                b.duration,
                b.date.strftime("%Y-%m-%d"),
                b.time.strftime("%H:%M") if b.time else "",
                status
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="Bookings.xlsx"'
        wb.save(response)

        return response



class ExportServiceRequestsExcel(APIView):
    permission_classes = [IsAuthenticated]  

    def get(self, request, *args, **kwargs):
        service_requests = ServiceEnquiry.objects.all().order_by("-created_at")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Service Requests"

        headers = ["Name", "Message", "Created At"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add rows
        for req in service_requests:
            ws.append([
                req.name,
                req.message,
                req.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="ServiceRequests.xlsx"'
        wb.save(response)

        return response
    
    
    
class ExportEnquiryBookingExcel(APIView):
    permission_classes = [IsAuthenticated] 

    def get(self, request, *args, **kwargs):
        enquiry_bookings = EnquiryBooking.objects.all().order_by("-created_at")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Service Requests"

        headers = ["Name", "Email","Title","Phone", "Number of Persons","Duration","Date", "Created At"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for req in enquiry_bookings:
            ws.append([
                req.name,
                req.email,
                req.title,
                req.phone,
                req.number_of_persons,
                req.duration,
                req.date.strftime("%Y-%m-%d"),
                req.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="ServiceRequests.xlsx"'
        wb.save(response)

        return response